from openpyxl import Workbook
from openpyxl import load_workbook

class ExcelImporter:

    def __init__(self, filePath):
        self.filePath=filePath
        self.excelFile = load_workbook(filePath)
        self.sheet=self.excelFile['Sheet1']

    def readData(self):

        wb = Workbook()

        return self.sheet['D5'].value
